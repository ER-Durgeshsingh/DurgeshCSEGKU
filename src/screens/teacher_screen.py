import streamlit as st

from src.ui.base_layout import style_background_dashboard, style_base_layout

from src.components.header import header_dashboard
from src.components.footer import footer_dashboard
from src.components.subject_card import subject_card
from src.database.db import (
    check_teacher_exists, create_teacher, teacher_login, get_teacher_subjects,
    get_attendance_for_teacher, get_teacher_by_username_or_email, create_teacher_otp,
    verify_teacher_otp, update_teacher_password
)
from src.components.dialog_create_subject import create_subject_dialog
from src.components.dialog_share_subject import share_subject_dialog
from src.components.dialog_add_photo import add_photos_dialog

from src.pipelines.face_pipeline import predict_attendance
from src.components.dialog_attendance_results import attendance_result_dialog
import numpy as np

from datetime import datetime
from io import BytesIO
import smtplib
from email.message import EmailMessage

import pandas as pd

try:
    from fpdf import FPDF
except Exception:
    FPDF = None

from src.database.config import supabase


from src.components.dialog_voice_attendance import voice_attendance_dialog
def teacher_screen():

    style_background_dashboard()
    style_base_layout()

    if "teacher_data" in st.session_state:
        teacher_dashboard()
    elif 'teacher_login_type' not in st.session_state or st.session_state.teacher_login_type=="login":
        teacher_screen_login()
    elif st.session_state.teacher_login_type == "register":
        teacher_screen_register()





def teacher_dashboard():
    teacher_data = st.session_state.teacher_data
    c1, c2 = st.columns(2, vertical_alignment='center', gap='xxlarge')
    with c1:
        header_dashboard()
    with c2:
        st.subheader(f"""Welcome, {teacher_data['name']} """)
        if st.button("Logout", type='secondary', key='teacher_logout_btn', shortcut="control+backspace"):
            st.session_state['is_logged_in'] = False
            del st.session_state.teacher_data 
            st.rerun()


    st.space()

    if "current_teacher_tab" not in st.session_state:
        st.session_state.current_teacher_tab = 'take_attendance'
    tab1, tab2, tab3, tab4 = st.columns(4)


    with tab1:
        type1 = "primary" if st.session_state.current_teacher_tab == 'take_attendance' else "tertiary"
        if st.button('Take Attendance', type=type1, width='stretch', icon=':material/ar_on_you:', key='teacher_tab_take_attendance_btn'):
            st.session_state.current_teacher_tab = 'take_attendance'
            st.rerun()

    with tab2:
        type2 = "primary" if st.session_state.current_teacher_tab == 'manage_subjects' else "tertiary"
        if st.button('Manage Subjects', type=type2, width='stretch', icon=':material/book_ribbon:', key='teacher_tab_manage_subjects_btn'):
            st.session_state.current_teacher_tab = 'manage_subjects'
            st.rerun()

    with tab3:
        type3 = "primary" if st.session_state.current_teacher_tab == 'attendance_records' else "tertiary"
        if st.button('Attendance Records', type=type3, width='stretch', icon=':material/cards_stack:', key='teacher_tab_attendance_records_btn'):
            st.session_state.current_teacher_tab = 'attendance_records'
            st.rerun()

    with tab4:
        type4 = "primary" if st.session_state.current_teacher_tab == 'admin_dashboard' else "tertiary"
        if st.button('Admin Dashboard', type=type4, width='stretch', icon=':material/dashboard:', key='teacher_tab_admin_dashboard_btn'):
            st.session_state.current_teacher_tab = 'admin_dashboard'
            st.rerun()


    st.divider()

    if st.session_state.current_teacher_tab == "take_attendance":
        teacher_tab_take_attendance()
    if st.session_state.current_teacher_tab == "manage_subjects":
        teacher_tab_manage_subjects()
    if st.session_state.current_teacher_tab == "attendance_records":
        teacher_tab_attendance_records()
    if st.session_state.current_teacher_tab == "admin_dashboard":
        teacher_tab_admin_dashboard()

    


    footer_dashboard()

def teacher_tab_take_attendance():
    teacher_id = st.session_state.teacher_data['teacher_id']
    st.header('Take AI Attendance')


    if 'attendance_images' not in st.session_state:
        st.session_state.attendance_images = []

    subjects = get_teacher_subjects(teacher_id)

    if not subjects:
        st.warning('You havent created any subjects yet! Please create one to begin!')
        return
    
    subject_options = {f"{s['name']} - {s['subject_code']}": s['subject_id'] for s in subjects}

    col1, col2 = st.columns([3,1], vertical_alignment='bottom')

    with col1:
        selected_subject_label = st.selectbox('Select Subject', options=list(subject_options.keys()), key='teacher_attendance_subject_select')

    with col2:
        if st.button('Add Photos', type='primary', icon=':material/photo_prints:', width='stretch', key='teacher_add_photos_btn'):
            add_photos_dialog()

    selected_subject_id = subject_options[selected_subject_label]

    st.divider()

    if st.session_state.attendance_images:
        st.header('Added Photos')
        gallery_cols = st.columns(4)

        for idx, img in enumerate(st.session_state.attendance_images):
            with gallery_cols[idx % 4 ]:
                st.image(img, width='stretch', caption=f'Photo {idx+1}')
    has_photos = bool(st.session_state.attendance_images)
    c1, c2, c3 = st.columns(3)

    with c1:
        if st.button('Clear all photos', width='stretch', type='tertiary', icon=':material/delete:', disabled=not has_photos, key='teacher_clear_photos_btn'):
            st.session_state.attendance_images = []
            st.rerun()


    with c2:
        
        if st.button('Run Face Analysis', width='stretch', type='secondary', icon=':material/analytics:', disabled=not has_photos, key='teacher_run_face_analysis_btn'):
            with st.spinner('Deep scanning classroom photos...'):
                all_detected_ids = {}

                for idx, img in enumerate(st.session_state.attendance_images):
                    img_np = np.array(img.convert('RGB'))
                    detected, _, _ = predict_attendance(img_np)


                    if detected:
                        for sid in detected.keys():
                            student_id = int(sid)

                            all_detected_ids.setdefault(student_id, []).append(f"Photo {idx+1}")

                enrolled_res = supabase.table('subject_students').select("*, students(*)").eq('subject_id',selected_subject_id ).execute()
                enrolled_students = enrolled_res.data

                if not enrolled_students:
                    st.warning('No students enrolled in this course')
                else:

                    results, attendance_to_log  = [], []

                    current_timestamp = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


                    for node in enrolled_students:
                        student = node['students']
                        sources = all_detected_ids.get(int(student['student_id']), [])
                        is_present= len(sources) > 0

                        results.append({
                            "Name": student['name'],
                            "ID": student['student_id'],
                            "Source": ", ".join(sources) if is_present else "-",
                            "Status": "✅ Present" if is_present else "❌ Absent"
                        })

                        attendance_to_log.append({
                            'student_id': student['student_id'],
                            'subject_id': selected_subject_id,
                            'timestamp': current_timestamp,
                            'is_present': bool(is_present)
                        })

                attendance_result_dialog(pd.DataFrame(results), attendance_to_log)

    with c3:
        if st.button('Use Voice Attendance', type='primary', width='stretch', icon=':material/mic:', key='teacher_use_voice_attendance_btn'):
            voice_attendance_dialog(selected_subject_id)











def teacher_tab_manage_subjects():
    teacher_id = st.session_state.teacher_data['teacher_id']
    col1, col2 = st.columns(2)
    with col1:
        st.header('Manage Subjects', width='stretch')

    with col2:
        if st.button('Create New Subject', width='stretch', key='teacher_create_subject_open_btn'):
            create_subject_dialog(teacher_id)

    subjects = get_teacher_subjects(teacher_id)
    if subjects:
        for sub in subjects:
            subject_id = sub.get('subject_id', sub.get('subject_code'))
            stats = [
                ("🫂", "Students", sub.get('total_students', 0)),
                ("🕰️", "Classes", sub.get('total_classes', 0)),
            ]

            def share_btn(sub=sub, subject_id=subject_id):
                if st.button(
                    f"Share Code: {sub['name']}",
                    key=f"share_subject_{subject_id}_{sub['subject_code']}",
                    icon=":material/share:",
                ):
                    share_subject_dialog(sub['name'], sub['subject_code'])
                st.space()

            subject_card(
                name=sub['name'],
                code=sub['subject_code'],
                section=sub['section'],
                stats=stats,
                footer_callback=share_btn,
            )
    else:
        st.info("NO SUBJECTS FOUND. CREATE ONE ABOVE")


def _parse_attendance_datetime(value):
    """Parse Supabase timestamp safely for display, filtering and PDF export."""
    if not value:
        return None
    try:
        value = str(value).replace("Z", "+00:00")
        return datetime.fromisoformat(value)
    except Exception:
        try:
            return datetime.strptime(str(value).split(".")[0], "%Y-%m-%dT%H:%M:%S")
        except Exception:
            return None


def _safe_pdf_text(value):
    """FPDF default fonts support latin text only, so keep PDF text safe."""
    return str(value).encode("latin-1", "replace").decode("latin-1")


def _make_monthly_attendance_pdf(monthly_df, summary_df, teacher_name, month_label):
    if FPDF is None:
        return None

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, _safe_pdf_text("Monthly Attendance Report"), ln=True, align="C")

    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 7, _safe_pdf_text(f"Teacher: {teacher_name}"), ln=True)
    pdf.cell(0, 7, _safe_pdf_text(f"Month: {month_label}"), ln=True)
    pdf.cell(0, 7, _safe_pdf_text(f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"), ln=True)
    pdf.ln(4)

    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, _safe_pdf_text("Summary"), ln=True)

    pdf.set_font("Arial", "B", 9)
    widths = [55, 32, 28, 28, 32]
    headers = ["Subject", "Subject Code", "Present", "Total", "Percentage"]

    for w, h in zip(widths, headers):
        pdf.cell(w, 8, _safe_pdf_text(h), border=1, align="C")
    pdf.ln()

    pdf.set_font("Arial", "", 9)
    for _, row in summary_df.iterrows():
        values = [
            row.get("Subject", ""),
            row.get("Subject Code", ""),
            row.get("Present_Count", 0),
            row.get("Total_Count", 0),
            row.get("Percentage", "0%"),
        ]

        for w, v in zip(widths, values):
            pdf.cell(w, 8, _safe_pdf_text(v)[:30], border=1, align="C")
        pdf.ln()

    pdf.ln(5)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, _safe_pdf_text("Detailed Attendance"), ln=True)

    pdf.set_font("Arial", "B", 8)
    widths = [28, 32, 22, 35, 34, 20]
    headers = ["Date/Time", "Student", "Student ID", "University Roll No", "Subject", "Status"]

    for w, h in zip(widths, headers):
        pdf.cell(w, 7, _safe_pdf_text(h), border=1, align="C")
    pdf.ln()

    pdf.set_font("Arial", "", 8)

    for _, row in monthly_df.sort_values(by="DateTime", ascending=False).iterrows():
        values = [
            row.get("Time", ""),
            row.get("Student", ""),
            row.get("Student ID", ""),
            row.get("University Roll Number", "N/A"),
            row.get("Subject", ""),
            row.get("Status", ""),
        ]

        for w, v in zip(widths, values):
            pdf.cell(w, 7, _safe_pdf_text(v)[:25], border=1)
        pdf.ln()

    output = pdf.output(dest="S")

    if isinstance(output, str):
        return output.encode("latin-1")

    return bytes(output)


def teacher_tab_attendance_records():
    st.header('Attendance Records')

    teacher_id = st.session_state.teacher_data['teacher_id']
    teacher_name = st.session_state.teacher_data.get('name', 'Teacher')

    records = get_attendance_for_teacher(teacher_id)

    if not records:
        st.info('No attendance records found yet.')
        return
    
    df = _attendance_dataframe(records)

    summary = (
        df.groupby(['ts_group', 'Time', 'Subject', 'Subject Code'])
        .agg(
            Present_Count=('is_present', 'sum'),
            Total_Count=('is_present', 'count')
        ).reset_index()
    )

    summary['Attendance Stats'] = (
        "Present " + summary['Present_Count'].astype(str) + " / "
        + summary['Total_Count'].astype(str) + ' Students'
    )

    display_df = (summary.sort_values(by='ts_group', ascending=False)
                  [['Time', 'Subject', 'Subject Code', 'Attendance Stats']])
    
    st.subheader('All Attendance Sessions')
    st.dataframe(display_df, width='stretch', hide_index=True)

    excel_bytes = _make_excel_bytes({
        "All Attendance": df.drop(columns=['DateTime'], errors='ignore'),
        "Session Summary": display_df,
    })
    st.download_button(
        "Download Attendance Excel",
        data=excel_bytes,
        file_name="attendance_records.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width='stretch',
        key='teacher_attendance_excel_download_btn'
    )

    st.divider()
    st.subheader('Download Month-wise Attendance PDF')

    month_df = df[df['Month'] != 'Unknown'].copy()
    if month_df.empty:
        st.warning('Valid monthly records are not available for PDF download.')
        return

    month_options = (
        month_df[['Month', 'Month Label']]
        .drop_duplicates()
        .sort_values(by='Month', ascending=False)
    )
    month_labels = month_options['Month Label'].tolist()
    month_key_map = dict(zip(month_options['Month Label'], month_options['Month']))

    selected_month_label = st.selectbox(
        'Select Month',
        options=month_labels,
        key='teacher_monthly_pdf_month_select'
    )

    selected_month = month_key_map[selected_month_label]
    monthly_df = month_df[month_df['Month'] == selected_month].copy()

    monthly_summary = (
        monthly_df.groupby(['Subject', 'Subject Code'])
        .agg(Present_Count=('is_present', 'sum'), Total_Count=('is_present', 'count'))
        .reset_index()
    )
    monthly_summary['Percentage'] = (
        (monthly_summary['Present_Count'] / monthly_summary['Total_Count']) * 100
    ).round(2).astype(str) + '%'

    st.caption(f"{selected_month_label} records: {len(monthly_df)} entries")
    st.dataframe(
        monthly_summary[['Subject', 'Subject Code', 'Present_Count', 'Total_Count', 'Percentage']],
        width='stretch',
        hide_index=True
    )

    pdf_bytes = _make_monthly_attendance_pdf(
        monthly_df=monthly_df,
        summary_df=monthly_summary,
        teacher_name=teacher_name,
        month_label=selected_month_label,
    )

    register_excel_bytes = _make_register_style_excel(
        monthly_df=monthly_df,
        teacher_name=teacher_name,
        month_label=selected_month_label
    )

    st.download_button(
        label=f"Download {selected_month_label} Register Excel",
        data=register_excel_bytes,
        file_name=f"attendance_register_{selected_month}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type='primary',
        width='stretch',
        key='teacher_register_excel_download_btn'
    )

    if pdf_bytes:
        st.download_button(
            label=f"Download {selected_month_label} PDF",
            data=pdf_bytes,
            file_name=f"attendance_report_{selected_month}.pdf",
            mime="application/pdf",
            type='primary',
            width='stretch',
            key='teacher_monthly_attendance_pdf_download_btn'
        )
    else:
        st.error('PDF package missing. Run: pip install fpdf2')



def _make_excel_bytes(df_map):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in df_map.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
    output.seek(0)
    return output.getvalue()


def _make_register_style_excel(monthly_df, teacher_name, month_label):
    output = BytesIO()
    df = monthly_df.copy()

    df = df[df["DateTime"].notna()].copy()
    if df.empty:
        return _make_excel_bytes({"Monthly Register": pd.DataFrame()})

    df["Day"] = df["DateTime"].dt.day

    register_df = (
        df.pivot_table(
            index=["Student", "Student ID", "University Roll Number"],
            columns="Day",
            values="Status",
            aggfunc=lambda x: "P" if "Present" in list(x) else "A"
        )
        .reset_index()
    )

    for day in range(1, 32):
        if day not in register_df.columns:
            register_df[day] = ""

    register_df = register_df[
        ["Student", "Student ID", "University Roll Number"] + list(range(1, 32))
    ]

    register_df.rename(columns={
        "Student": "Student Name",
        "Student ID": "Student ID",
        "University Roll Number": "University Roll No"
    }, inplace=True)

    subject_name = monthly_df["Subject"].iloc[0] if not monthly_df.empty else "N/A"
    subject_code = monthly_df["Subject Code"].iloc[0] if not monthly_df.empty else "N/A"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        register_df.to_excel(
            writer,
            sheet_name="Monthly Register",
            index=False,
            startrow=6
        )

        ws = writer.book["Monthly Register"]
        ws.merge_cells("A1:AH1")
        ws["A1"] = "ATTENDANCE REGISTER"
        ws["A2"] = f"Teacher: {teacher_name}"
        ws["A3"] = f"Subject: {subject_name}"
        ws["A4"] = f"Subject Code: {subject_code}"
        ws["A5"] = f"Month: {month_label}"

        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[7]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 22
        for col in range(4, 35):
            ws.column_dimensions[get_column_letter(col)].width = 4

    output.seek(0)
    return output.getvalue()


def _send_gmail(to_email, subject, body):
    try:
        gmail_address = st.secrets["GMAIL_ADDRESS"]
        gmail_app_password = st.secrets["GMAIL_APP_PASSWORD"]

        msg = EmailMessage()
        msg["From"] = gmail_address
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.set_content(body)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(gmail_address, gmail_app_password)
            smtp.send_message(msg)

        return True, "OTP sent successfully"

    except Exception as e:
        return False, str(e)



def _send_teacher_otp(identifier, purpose):
    teacher = get_teacher_by_username_or_email(identifier)
    if not teacher:
        return None, False, "Teacher not found."
    email = teacher.get("email")
    if not email:
        return teacher, False, "This teacher account has no Gmail/email saved. Register again with email or update database."
    otp = create_teacher_otp(teacher["teacher_id"], purpose=purpose)
    subject = "CSEClass OTP Code"
    body = f"Your OTP for {purpose.replace('_', ' ')} is: {otp}\n\nThis OTP is valid for 10 minutes."
    sent, msg = _send_gmail(email, subject, body)
    if not sent:
        # Local/dev fallback so project remains testable even before Gmail SMTP setup.
        st.session_state[f"dev_otp_{purpose}_{teacher['teacher_id']}"] = otp
        return teacher, False, msg + f"\nDEV OTP: {otp}"
    return teacher, True, msg


def _attendance_dataframe(records):
    data = []

    for r in records:
        ts = r.get('timestamp')
        dt = _parse_attendance_datetime(ts)

        student = r.get('students') or {}
        subject = r.get('subjects') or {}
        is_present = bool(r.get('is_present', False))

        data.append({
            "ts_group": dt.strftime("%Y-%m-%d %H:%M:%S") if dt else str(ts),
            "DateTime": dt,
            "Month": dt.strftime("%Y-%m") if dt else "Unknown",
            "Month Label": dt.strftime("%B %Y") if dt else "Unknown",
            "Time": dt.strftime("%Y-%m-%d %I:%M %p") if dt else "N/A",

            "Subject": subject.get('name', 'N/A'),
            "Subject Code": subject.get('subject_code', 'N/A'),

            "Student": student.get('name', 'N/A'),
            "Student ID": r.get(
                'student_id',
                student.get('student_id', 'N/A')
            ),

            "University Roll Number": student.get(
                'university_roll_number',
                'N/A'
            ),

            "is_present": is_present,
            "Status": "Present" if is_present else "Absent",
        })

    return pd.DataFrame(data)


def teacher_tab_admin_dashboard():
    st.header("Admin Dashboard & Auto Attendance Analytics")

    teacher_id = st.session_state.teacher_data['teacher_id']
    subjects = get_teacher_subjects(teacher_id)
    records = get_attendance_for_teacher(teacher_id)
    df = _attendance_dataframe(records)

    total_subjects = len(subjects)
    total_students = sum(int(s.get('total_students', 0)) for s in subjects)
    total_sessions = df['ts_group'].nunique() if not df.empty else 0
    present_rate = round((df['is_present'].sum() / len(df)) * 100, 2) if not df.empty else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Subjects", total_subjects)
    c2.metric("Enrolled Students", total_students)
    c3.metric("Attendance Sessions", total_sessions)
    c4.metric("Overall Present %", f"{present_rate}%")

    if df.empty:
        st.info("No attendance analytics available yet. Take attendance first.")
        return

    st.subheader("Subject-wise Analytics")

    subject_summary = (
        df.groupby(['Subject', 'Subject Code'])
        .agg(Present=('is_present', 'sum'), Total=('is_present', 'count'))
        .reset_index()
    )

    subject_summary['Percentage'] = (
        (subject_summary['Present'] / subject_summary['Total']) * 100
    ).round(2)

    st.dataframe(subject_summary, width='stretch', hide_index=True)
    st.bar_chart(subject_summary.set_index('Subject')['Percentage'])

    st.subheader("Student-wise Low Attendance Alerts")

    student_summary = (
        df.groupby(['Student', 'Student ID', 'University Roll Number', 'Subject'])
        .agg(Present=('is_present', 'sum'), Total=('is_present', 'count'))
        .reset_index()
    )

    student_summary['Percentage'] = (
        (student_summary['Present'] / student_summary['Total']) * 100
    ).round(2)

    low_df = student_summary[student_summary['Percentage'] < 75].sort_values('Percentage')

    if low_df.empty:
        st.success("No low-attendance students below 75%.")
    else:
        st.warning("Students below 75% attendance")
        st.dataframe(low_df, width='stretch', hide_index=True)

    st.subheader("Month-wise Trend")

    month_summary = (
        df[df['Month'] != 'Unknown'].groupby('Month')
        .agg(Present=('is_present', 'sum'), Total=('is_present', 'count'))
        .reset_index()
    )

    month_summary['Percentage'] = (
        (month_summary['Present'] / month_summary['Total']) * 100
    ).round(2)

    st.line_chart(month_summary.set_index('Month')['Percentage'])

    excel_bytes = _make_excel_bytes({
        "Subject Analytics": subject_summary,
        "Student Analytics": student_summary,
        "Low Attendance": low_df,
        "Monthly Trend": month_summary,
        "Raw Attendance": df.drop(columns=['DateTime'], errors='ignore'),
    })

    st.download_button(
        "Download Full Analytics Excel",
        data=excel_bytes,
        file_name="attendance_analytics.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        width="stretch",
        key="download_admin_analytics_excel_btn",
    )

def login_teacher(username, password):
    if not username or not password:
        return False
    
    teacher = teacher_login(username, password)

    if teacher:
        st.session_state.user_role ='teacher'
        st.session_state.teacher_data = teacher
        st.session_state.is_logged_in = True
        return True
    

    return False
def teacher_screen_login():
    c1, c2 = st.columns(2, vertical_alignment='center', gap='xxlarge')
    with c1:
        header_dashboard()
    with c2:
        if st.button("Go back to Home", type='secondary', key='teacher_back_home_btn', shortcut="control+backspace"):
            st.session_state['login_type'] = None
            st.rerun()

    st.header('Teacher Login', text_alignment='center')
    login_tab, otp_tab, forgot_tab = st.tabs(["Password Login", "Gmail OTP Login", "Forgot Password"])

    with login_tab:
        teacher_username = st.text_input("Enter username or Gmail", placeholder='durgeshsingh or gmail@gmail.com', key='teacher_login_username')
        teacher_pass = st.text_input("Enter password", type='password', placeholder="Enter password", key='teacher_login_password')
        st.divider()
        btnc1, btnc2 = st.columns(2)
        with btnc1:
            if st.button('Login', icon=':material/passkey:', shortcut='control+enter', width='stretch', key='teacher_login_submit_btn'):
                if login_teacher(teacher_username, teacher_pass):
                    st.toast("welcome back!", icon="👋")
                    import time
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("Invalid username/email and password combo")
        with btnc2:
            if st.button('Register Instead', type="primary", icon=':material/person_add:', width='stretch', key='teacher_switch_register_btn'):
                st.session_state.teacher_login_type = 'register'
                st.rerun()

    with otp_tab:
        identifier = st.text_input("Enter username or Gmail", key="teacher_otp_identifier", placeholder="teacher username or Gmail")
        if st.button("Send OTP on Gmail", type="primary", width="stretch", key="teacher_send_login_otp_btn"):
            teacher, sent, msg = _send_teacher_otp(identifier, purpose="login")
            if teacher:
                st.session_state.login_otp_teacher_id = teacher['teacher_id']
                st.session_state.login_otp_teacher = teacher
            (st.success if sent else st.warning)(msg)
        otp_code = st.text_input("Enter OTP", key="teacher_login_otp_code")
        if st.button("Verify OTP & Login", width="stretch", key="teacher_verify_login_otp_btn"):
            teacher = st.session_state.get('login_otp_teacher')
            if teacher and verify_teacher_otp(teacher['teacher_id'], otp_code, purpose="login"):
                st.session_state.user_role = 'teacher'
                st.session_state.teacher_data = teacher
                st.session_state.is_logged_in = True
                st.success("OTP verified. Login successful.")
                st.rerun()
            else:
                st.error("Invalid or expired OTP.")

    with forgot_tab:
        fp_identifier = st.text_input("Enter your registered username or Gmail", key="forgot_password_identifier")
        if st.button("Send Password Reset OTP", type="primary", width="stretch", key="teacher_send_reset_otp_btn"):
            teacher, sent, msg = _send_teacher_otp(fp_identifier, purpose="reset_password")
            if teacher:
                st.session_state.reset_otp_teacher_id = teacher['teacher_id']
            (st.success if sent else st.warning)(msg)
        reset_otp = st.text_input("Enter Reset OTP", key="teacher_reset_otp_code")
        new_pass = st.text_input("New Password", type="password", key="teacher_new_password")
        new_pass2 = st.text_input("Confirm New Password", type="password", key="teacher_new_password_confirm")
        if st.button("Reset Password", width="stretch", key="teacher_reset_password_btn"):
            teacher_id = st.session_state.get('reset_otp_teacher_id')
            if not teacher_id:
                st.error("Please send OTP first.")
            elif new_pass != new_pass2:
                st.error("Password does not match.")
            elif len(new_pass) < 6:
                st.error("Password should be at least 6 characters.")
            elif verify_teacher_otp(teacher_id, reset_otp, purpose="reset_password"):
                update_teacher_password(teacher_id, new_pass)
                st.success("Password reset successful. Now login with new password.")
            else:
                st.error("Invalid or expired OTP.")

    footer_dashboard()


def register_teacher(teacher_username, teacher_name, teacher_email, teacher_pass, teacher_pass_confirm):
    if not teacher_username or not teacher_name or not teacher_email or not teacher_pass:
        return False, "All Fields are required!"
    if check_teacher_exists(teacher_username):
        return False, "Username already taken"
    if teacher_pass != teacher_pass_confirm:
        return False, "Password doesn't match"
    
    try:
        create_teacher(teacher_username, teacher_pass, teacher_name, teacher_email)
        return True, "Sucessfully Created! Login Now"
    except Exception as e:
        return False, "Unexpected Error!"
    

def teacher_screen_register():
    c1, c2 = st.columns(2, vertical_alignment='center', gap='xxlarge')
    with c1:
        header_dashboard()
    with c2:
        if st.button("Go back to Home", type='secondary', key='teacher_register_back_home_btn', shortcut="control+backspace"):
            st.session_state['login_type'] = None
            st.rerun()



    st.header('Register your teacher profile')

    st.space()
    st.space()

    
    teacher_username = st.text_input("Enter username", placeholder='durgeshsingh', key='teacher_register_username')

    teacher_name = st.text_input("Enter name", placeholder='Durgeshwar kumar Singh', key='teacher_register_name')

    teacher_email = st.text_input("Enter Gmail / Email", placeholder='yourgmail@gmail.com', key='teacher_register_email')

    teacher_pass = st.text_input("Enter password", type='password', placeholder="Enter password", key='teacher_register_password')

    teacher_pass_confirm = st.text_input("Confirm your password", type='password', placeholder="Enter password", key='teacher_register_confirm_password')

    st.divider()

    btnc1, btnc2 = st.columns(2)

    with btnc1:
        if st.button('Register now', icon=':material/passkey:', shortcut='control+enter', width='stretch', key='teacher_register_submit_btn'):
            success, message = register_teacher(teacher_username, teacher_name, teacher_email, teacher_pass, teacher_pass_confirm)
            if success:
                st.success(message)
                import time
                time.sleep(2)
                st.session_state.teacher_login_type = "login"
                st.rerun()
            else:
                st.error(message)


    with btnc2:
        if st.button('Login Instead', type="primary", icon=':material/passkey:', width='stretch', key='teacher_switch_login_btn'):
            st.session_state.teacher_login_type = 'login'

    footer_dashboard()
